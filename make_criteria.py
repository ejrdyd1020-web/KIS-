from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

wb = Workbook()

BLUE_HEADER   = '1F4E79'
BREAKOUT_HDR  = '2E75B6'
REVERSION_HDR = '375623'
COMMON_HDR    = '7030A0'
ROW_EVEN      = 'DCE6F1'
ROW_ODD       = 'FFFFFF'
SCORE_BG      = 'FFF2CC'
WHITE         = 'FFFFFF'

def fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)

def border():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def left_align():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def cell_font(bold=False, sz=10, color='000000'):
    return Font(name='Arial', bold=bold, size=sz, color=color)

def write_title(ws, text, bg):
    ws.row_dimensions[1].height = 36
    ws.merge_cells('A1:D1')
    c = ws['A1']
    c.value     = text
    c.font      = Font(name='Arial', bold=True, size=14, color=WHITE)
    c.fill      = fill(bg)
    c.alignment = center()
    c.border    = border()

def write_section_header(ws, row, text, bg):
    ws.row_dimensions[row].height = 20
    ws.merge_cells(f'A{row}:D{row}')
    c = ws[f'A{row}']
    c.value     = text
    c.font      = Font(name='Arial', bold=True, size=10, color=WHITE)
    c.fill      = fill(bg)
    c.alignment = center()
    c.border    = border()

def write_col_header(ws, row, hdr_bg, labels=('No', '항목', '기준', '비고')):
    ws.row_dimensions[row].height = 22
    for col, txt in enumerate(labels, 1):
        c = ws.cell(row=row, column=col, value=txt)
        c.font      = Font(name='Arial', bold=True, color=WHITE, size=11)
        c.fill      = fill(hdr_bg)
        c.alignment = center()
        c.border    = border()

def write_data_row(ws, row, values, bg):
    ws.row_dimensions[row].height = 18
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill      = fill(bg)
        c.font      = cell_font()
        c.border    = border()
        c.alignment = left_align() if col == 2 else center()

def set_col_widths(ws, widths):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════
# 시트 1: BREAKOUT
# ══════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'BREAKOUT 전략'
ws1.sheet_view.showGridLines = False
set_col_widths(ws1, [5, 28, 22, 30])

write_title(ws1, '전략 A : BREAKOUT  (09:00 ~ 09:10)', BLUE_HEADER)

BSEC = '41719C'
r = 3
write_section_header(ws1, r, '[ 매수 조건 ]', BSEC)
r += 1
write_col_header(ws1, r, BREAKOUT_HDR)

buy_rows = [
    ('운용 시간',     '09:00 ~ 09:10',        '장 초반 10분'),
    ('등락률',        '+3.0% ~ +25.0%',       '갭 상승 기준 완화'),
    ('전일 고가 돌파','현재가 > 전일 고가',    '데이터 없으면 시가 대비 +1%'),
    ('거래량 급증',   '전일 대비 3배 이상',   '분당 환산'),
    ('거래대금 급증', '전일 대비 2배 이상',   '분당 환산'),
    ('체결강도',      '100% 이상',             ''),
    ('가격 범위',     '1,000원 이상',          '동전주 제외'),
    ('제외 종목',     'ETF/ETN/리츠/스팩/우선주', '종목명 키워드 필터'),
    ('스캔 주기',     '5초',                   '실시간'),
]
for i, (item, crit, note) in enumerate(buy_rows):
    r += 1
    bg = ROW_EVEN if i % 2 == 0 else ROW_ODD
    write_data_row(ws1, r, [i+1, item, crit, note], bg)

r += 2
write_section_header(ws1, r, '[ 점수 산정 — 70점 이상 매수 집행 ]', BSEC)
r += 1
write_col_header(ws1, r, BREAKOUT_HDR, ('No', '항목', '가중치', '만점 기준'))

score_rows = [
    ('체결강도',      '30%', '120% 이상 = 만점 30점'),
    ('등락률',        '30%', '15% 이상 = 만점 30점'),
    ('거래량 순위',   '20%', '1위=20점, 300위=0점'),
    ('거래대금 순위', '20%', '1위=20점, 300위=0점'),
]
for i, (item, w, note) in enumerate(score_rows):
    r += 1
    bg = SCORE_BG if i % 2 == 0 else 'FFFDE7'
    write_data_row(ws1, r, [i+1, item, w, note], bg)

r += 2
write_section_header(ws1, r, '[ 청산 조건 ]', BSEC)
r += 1
write_col_header(ws1, r, BREAKOUT_HDR)

exit_rows = [
    ('사전경고 (지정가 예약)', '-2.0%',            '도달 시 hard_stop 가격 지정가 매도 예약'),
    ('고정 손절',              '-3.0%',            '지정가 체결 실패 시 시장가 백업'),
    ('고정 익절',              '+5.0%',            ''),
    ('트레일링 스탑 작동',     '수익 +3.0% 이상',  '이후 고점 추적 시작'),
    ('트레일링 청산',          '고점 대비 -2.0%',  ''),
    ('MA120 이탈 손절',        '1분봉 MA120 하회', '시장가 즉시 청산'),
    ('장마감 강제청산',        '15:20',            ''),
]
for i, (item, crit, note) in enumerate(exit_rows):
    r += 1
    bg = 'FDECEA' if i % 2 == 0 else 'FFF5F5'
    write_data_row(ws1, r, [i+1, item, crit, note], bg)

# ══════════════════════════════════════════════════════════
# 시트 2: REVERSION
# ══════════════════════════════════════════════════════════
ws2 = wb.create_sheet('REVERSION 전략')
ws2.sheet_view.showGridLines = False
set_col_widths(ws2, [5, 30, 24, 30])

write_title(ws2, '전략 B : REVERSION  (09:10 ~ 15:20)', REVERSION_HDR)

RSEC = '70AD47'
r = 3
write_section_header(ws2, r, '[ 매수 조건 ]', RSEC)
r += 1
write_col_header(ws2, r, '538135')

rev_buy = [
    ('운용 시간',        '09:10 ~ 15:20',              ''),
    ('등락률',           '+1.5% ~ +20.0%',             ''),
    ('전일 거래대금',    '150억 이상',                  ''),
    ('거래량 급증',      '전일 대비 1.3배 이상',        ''),
    ('시가총액',         '100억 이상',                  ''),
    ('1분봉 스토캐스틱', 'K<20 -> K>D 골든크로스 (3봉 이내)', '침체권 탈출'),
    ('5분봉 MA20',       '현재가 >= 5분봉 MA20',        '상승 추세 확인'),
    ('5분봉 스토캐스틱', 'K < 50',                      '상위 TF 과열 아님'),
    ('체결강도',         '100% 이상',                   ''),
    ('스캔 주기',        '30초',                        ''),
]
for i, (item, crit, note) in enumerate(rev_buy):
    r += 1
    bg = 'E2EFDA' if i % 2 == 0 else ROW_ODD
    write_data_row(ws2, r, [i+1, item, crit, note], bg)

r += 2
write_section_header(ws2, r, '[ 종목 우선순위 배분 ]', RSEC)
r += 1
write_col_header(ws2, r, '538135', ('No', '항목', '비중', '비고'))
for i, (item, w, note) in enumerate([('거래대금', '70%', ''), ('거래량 배율', '30%', '')]):
    r += 1
    bg = SCORE_BG if i % 2 == 0 else 'FFFDE7'
    write_data_row(ws2, r, [i+1, item, w, note], bg)

r += 2
write_section_header(ws2, r, '[ 청산 조건 ]', RSEC)
r += 1
write_col_header(ws2, r, '538135')

rev_exit = [
    ('사전경고 (지정가 예약)', '-1.0%',               '도달 시 hard_stop 가격 지정가 매도 예약'),
    ('고정 손절',              '-1.5%',               '지정가 체결 실패 시 시장가 백업'),
    ('고정 익절',              '+3.0%',               '짧은 순환매'),
    ('트레일링 스탑 작동',     '수익 +2.0% 이상',     '이후 고점 추적 시작'),
    ('트레일링 청산',          '고점 대비 -1.5%',     ''),
    ('스토캐스틱 매도',        '과열권(80이상) 데드크로스', ''),
    ('MA120 이탈 손절',        '1분봉 MA120 하회',    '시장가 즉시 청산'),
    ('장마감 강제청산',        '15:20',               ''),
]
for i, (item, crit, note) in enumerate(rev_exit):
    r += 1
    bg = 'FDECEA' if i % 2 == 0 else 'FFF5F5'
    write_data_row(ws2, r, [i+1, item, crit, note], bg)

# ══════════════════════════════════════════════════════════
# 시트 3: 공통 설정
# ══════════════════════════════════════════════════════════
ws3 = wb.create_sheet('공통 설정')
ws3.sheet_view.showGridLines = False
set_col_widths(ws3, [5, 30, 24, 30])

write_title(ws3, '공통 설정', COMMON_HDR)

CSEC = '5B2C8A'
r = 3
write_section_header(ws3, r, '[ 자금 / 리스크 ]', CSEC)
r += 1
write_col_header(ws3, r, CSEC)

fund_rows = [
    ('총 운용 예산',         '10,000,000원',           ''),
    ('최대 동시 보유',       '6종목',                  '전략별 3종목'),
    ('일일 최대 손실 한도',  '총 예산의 5% (500,000원)', '초과 시 자동 중단'),
    ('포지션 모니터링',      '5초 간격',               '사전경고 진입 시 1초'),
    ('거래비용',             '수수료 0.03% + 거래세 0.18%', '합계 약 0.21%'),
]
for i, (item, crit, note) in enumerate(fund_rows):
    r += 1
    bg = 'EDE7F6' if i % 2 == 0 else ROW_ODD
    write_data_row(ws3, r, [i+1, item, crit, note], bg)

r += 2
write_section_header(ws3, r, '[ 시장 국면별 자금 배분 ]', CSEC)
r += 1
write_col_header(ws3, r, CSEC, ('No', '국면', 'BREAKOUT', 'REVERSION'))

phase_rows = [
    ('BULL (지수 MA5>MA20, +0.5% 이상)', '80%', '20%'),
    ('BEAR (그 외)',                      '20%', '80%'),
    ('판단 기준',                         '코스피 + 코스닥 평균', ''),
]
for i, (item, b, rv) in enumerate(phase_rows):
    r += 1
    bg = 'EDE7F6' if i % 2 == 0 else ROW_ODD
    write_data_row(ws3, r, [i+1, item, b, rv], bg)

os.makedirs('docs', exist_ok=True)
out = 'docs/매매기준표.xlsx'
wb.save(out)
print('저장 완료:', out)
